from __future__ import annotations

import argparse
import csv
import re
import sys
from pathlib import Path
from typing import Any, List, Tuple

from docx import Document
from docx.table import Table
from openpyxl import load_workbook
import xlrd


# ----------------------------
# Utilities
# ----------------------------
def is_empty(v: Any) -> bool:
    if v is None:
        return True
    if isinstance(v, str) and v.strip() == "":
        return True
    return False


def normalize_cell(v: Any) -> str:
    if v is None:
        return ""
    if hasattr(v, "isoformat"):
        try:
            return v.isoformat(sep=" ")
        except TypeError:
            return str(v)
    return str(v)


def sheet_used_bounds(values: List[List[Any]]) -> Tuple[int, int, int, int]:
    """
    values: 2D list [row][col] (0-based)
    Returns: (r0, r1, c0, c1) inclusive bounds of non-empty cells.
    If all empty -> (0, -1, 0, -1)
    """
    non_empty = []
    for r, row in enumerate(values):
        for c, v in enumerate(row):
            if is_empty(v):
                continue
            non_empty.append((r, c))
    if not non_empty:
        return (0, -1, 0, -1)
    rs = [rc[0] for rc in non_empty]
    cs = [rc[1] for rc in non_empty]
    return (min(rs), max(rs), min(cs), max(cs))


def safe_filename(name: str, max_len: int = 80) -> str:
    name = re.sub(r'[\\/:*?"<>|\s.。、，；：！？（）［］｛｝「」『』【】〈〉《》〔〕〖〗〘〙〚〛．]', "_", name)
    name = re.sub(r"_+", "_", name).strip("_")
    if not name:
        name = "sheet"
    return name[:max_len]


def write_csv(rows: List[List[str]], out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerows(rows)


# ----------------------------
# Sheet -> Matrix (fill merged)
# ----------------------------
def read_sheet_matrix_fill_merged(ws) -> List[List[str]]:
    max_row = ws.max_row or 1
    max_col = ws.max_column or 1

    mat: List[List[Any]] = []
    for r in range(1, max_row + 1):
        row_vals = []
        for c in range(1, max_col + 1):
            row_vals.append(ws.cell(row=r, column=c).value)
        mat.append(row_vals)

    for merged in ws.merged_cells.ranges:
        min_row, min_col, max_row2, max_col2 = merged.min_row, merged.min_col, merged.max_row, merged.max_col
        top_left = mat[min_row - 1][min_col - 1]
        for rr in range(min_row - 1, max_row2):
            for cc in range(min_col - 1, max_col2):
                mat[rr][cc] = top_left

    r0, r1, c0, c1 = sheet_used_bounds(mat)
    if r1 < r0 or c1 < c0:
        return []

    trimmed: List[List[str]] = []
    for r in range(r0, r1 + 1):
        trimmed.append([normalize_cell(mat[r][c]) for c in range(c0, c1 + 1)])
    return trimmed


# ----------------------------
# Matrix helpers
# ----------------------------
def split_blocks_by_blank_rows(matrix: List[List[str]], blank_rows: int = 1) -> List[List[List[str]]]:
    if not matrix:
        return []

    def row_is_blank(row: List[str]) -> bool:
        return all((cell.strip() == "") for cell in row)

    blocks: List[List[List[str]]] = []
    current: List[List[str]] = []
    blank_run = 0

    for row in matrix:
        if row_is_blank(row):
            blank_run += 1
            if current and blank_run >= blank_rows:
                blocks.append(current)
                current = []
        else:
            blank_run = 0
            current.append(row)

    if current:
        blocks.append(current)
    return blocks


def trim_empty_cols(block: List[List[str]]) -> List[List[str]]:
    if not block:
        return block

    ncols = max(len(r) for r in block)
    padded = [r + [""] * (ncols - len(r)) for r in block]

    non_empty_cols = []
    for c in range(ncols):
        col_has = any(padded[r][c].strip() != "" for r in range(len(padded)))
        if col_has:
            non_empty_cols.append(c)

    if not non_empty_cols:
        return []

    c0, c1 = min(non_empty_cols), max(non_empty_cols)
    return [row[c0 : c1 + 1] for row in padded]


# ----------------------------
# DOCX processing functions
# ----------------------------
def extract_table_from_docx_table(table: Table) -> List[List[str]]:
    rows = []
    for row in table.rows:
        cells = []
        for cell in row.cells:
            cell_text = cell.text.strip().replace("\n", " ").replace("\r", " ")
            cells.append(cell_text)
        rows.append(cells)
    return rows


def convert_docx_to_csv(docx_path: Path, out_dir: Path, blank_rows: int = 1) -> List[Path]:
    doc = Document(docx_path)
    out_dir.mkdir(parents=True, exist_ok=True)

    created: List[Path] = []
    base = safe_filename(docx_path.stem)

    table_count = 0

    for table in doc.tables:
        table_count += 1
        table_data = extract_table_from_docx_table(table)
        blocks = split_blocks_by_blank_rows(table_data, blank_rows=max(1, blank_rows))

        if not blocks:
            out_path = out_dir / f"{base}__table{table_count:02d}.csv"
            write_csv([["この表は空、または値が見つかりませんでした"]], out_path)
            created.append(out_path)
            continue

        if len(blocks) == 1:
            rows = trim_empty_cols(blocks[0])
            out_path = out_dir / f"{base}__table{table_count:02d}.csv"
            write_csv(rows if rows else [["表が生成できませんでした"]], out_path)
            created.append(out_path)
            continue

        for sub_no, block in enumerate(blocks, start=1):
            rows = trim_empty_cols(block)
            if not rows:
                continue
            out_path = out_dir / f"{base}__table{table_count:02d}_{sub_no:02d}.csv"
            write_csv(rows, out_path)
            created.append(out_path)

    if table_count == 0:
        out_path = out_dir / f"{base}.csv"
        write_csv([["このドキュメントに表は見つかりませんでした"]], out_path)
        created.append(out_path)

    return created


# ----------------------------
# XLSX processing functions
# ----------------------------
def convert_excel_to_csv_per_sheet(xlsx_path: Path, out_dir: Path, blank_rows: int = 1) -> List[Path]:
    wb = load_workbook(xlsx_path, data_only=True)
    out_dir.mkdir(parents=True, exist_ok=True)

    created: List[Path] = []
    base = safe_filename(xlsx_path.stem)

    for idx, sheet_name in enumerate(wb.sheetnames, start=1):
        ws = wb[sheet_name]
        matrix = read_sheet_matrix_fill_merged(ws)

        safe_sheet = safe_filename(sheet_name)

        if not matrix:
            out_path = out_dir / f"{base}__{idx:02d}_{safe_sheet}.csv"
            write_csv([["このシートは空、または値が見つかりませんでした"]], out_path)
            created.append(out_path)
            continue

        blocks = split_blocks_by_blank_rows(matrix, blank_rows=max(1, blank_rows))

        if len(blocks) == 1:
            rows = trim_empty_cols(blocks[0])
            out_path = out_dir / f"{base}__{idx:02d}_{safe_sheet}.csv"
            write_csv(rows if rows else [["表が生成できませんでした"]], out_path)
            created.append(out_path)
            continue

        table_no = 0
        for block in blocks:
            rows = trim_empty_cols(block)
            if not rows:
                continue
            table_no += 1
            out_path = out_dir / f"{base}__{idx:02d}_{safe_sheet}__table{table_no:02d}.csv"
            write_csv(rows, out_path)
            created.append(out_path)

        if table_no == 0:
            out_path = out_dir / f"{base}__{idx:02d}_{safe_sheet}.csv"
            write_csv([["表が生成できませんでした"]], out_path)
            created.append(out_path)

    return created


# ----------------------------
# XLS processing functions
# ----------------------------
def read_xls_matrix_fill_merged(ws) -> List[List[str]]:
    max_row = ws.nrows
    max_col = ws.ncols

    mat: List[List[Any]] = []
    for r in range(max_row):
        row_vals = []
        for c in range(max_col):
            row_vals.append(ws.cell_value(r, c))
        mat.append(row_vals)

    for crange in ws.merged_cells:
        rlo, rhi, clo, chi = crange
        top_left = mat[rlo][clo]
        for rr in range(rlo, rhi):
            for cc in range(clo, chi):
                mat[rr][cc] = top_left

    r0, r1, c0, c1 = sheet_used_bounds(mat)
    if r1 < r0 or c1 < c0:
        return []

    trimmed: List[List[str]] = []
    for r in range(r0, r1 + 1):
        trimmed.append([normalize_cell(mat[r][c]) for c in range(c0, c1 + 1)])
    return trimmed


def convert_xls_to_csv_per_sheet(xls_path: Path, out_dir: Path, blank_rows: int = 1) -> List[Path]:
    try:
        wb = xlrd.open_workbook(str(xls_path), formatting_info=True)
    except Exception:
        wb = xlrd.open_workbook(str(xls_path))

    out_dir.mkdir(parents=True, exist_ok=True)

    created: List[Path] = []
    base = safe_filename(xls_path.stem)

    for idx, ws in enumerate(wb.sheets(), start=1):
        sheet_name = ws.name
        matrix = read_xls_matrix_fill_merged(ws)

        safe_sheet = safe_filename(sheet_name)

        if not matrix:
            out_path = out_dir / f"{base}__{idx:02d}_{safe_sheet}.csv"
            write_csv([["このシートは空、または値が見つかりませんでした"]], out_path)
            created.append(out_path)
            continue

        blocks = split_blocks_by_blank_rows(matrix, blank_rows=max(1, blank_rows))

        if len(blocks) == 1:
            rows = trim_empty_cols(blocks[0])
            out_path = out_dir / f"{base}__{idx:02d}_{safe_sheet}.csv"
            write_csv(rows if rows else [["表が生成できませんでした"]], out_path)
            created.append(out_path)
            continue

        table_no = 0
        for block in blocks:
            rows = trim_empty_cols(block)
            if not rows:
                continue
            table_no += 1
            out_path = out_dir / f"{base}__{idx:02d}_{safe_sheet}__table{table_no:02d}.csv"
            write_csv(rows, out_path)
            created.append(out_path)

        if table_no == 0:
            out_path = out_dir / f"{base}__{idx:02d}_{safe_sheet}.csv"
            write_csv([["表が生成できませんでした"]], out_path)
            created.append(out_path)

    return created


# ----------------------------
# Directory traversal
# ----------------------------
def convert_tree(input_root: Path, output_root: Path, blank_rows: int = 1) -> Tuple[int, int]:
    """
    Convert all .xlsx, .xls and .docx files under input_root into output_root,
    preserving subdirectory structure.
    Returns: (ok_count, ng_count) for documents.
    """
    ok = 0
    ng = 0

    xlsx_files = sorted(input_root.rglob("*.xlsx"))
    for xlsx in xlsx_files:
        if xlsx.name.startswith("~$"):
            continue

        rel_dir = xlsx.parent.relative_to(input_root)
        out_dir = output_root / rel_dir

        try:
            convert_excel_to_csv_per_sheet(xlsx, out_dir, blank_rows=blank_rows)
            ok += 1
            print(f"[OK] {xlsx} -> {out_dir}")
        except Exception as e:
            ng += 1
            print(f"[NG] {xlsx} : {e}", file=sys.stderr)

    xls_files = sorted(input_root.rglob("*.xls"))
    for xls in xls_files:
        if xls.name.startswith("~$"):
            continue

        rel_dir = xls.parent.relative_to(input_root)
        out_dir = output_root / rel_dir

        try:
            convert_xls_to_csv_per_sheet(xls, out_dir, blank_rows=blank_rows)
            ok += 1
            print(f"[OK] {xls} -> {out_dir}")
        except Exception as e:
            ng += 1
            print(f"[NG] {xls} : {e}", file=sys.stderr)

    docx_files = sorted(input_root.rglob("*.docx"))
    for docx in docx_files:
        if docx.name.startswith("~$"):
            continue

        rel_dir = docx.parent.relative_to(input_root)
        out_dir = output_root / rel_dir

        try:
            convert_docx_to_csv(docx, out_dir, blank_rows=blank_rows)
            ok += 1
            print(f"[OK] {docx} -> {out_dir}")
        except Exception as e:
            ng += 1
            print(f"[NG] {docx} : {e}", file=sys.stderr)

    return ok, ng


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--in", dest="in_dir", default="設計書", help='Input root directory (default: "設計書")')
    ap.add_argument("--out", dest="out_dir", default="converted_csv_design", help='Output root directory (default: "converted_csv_design")')
    ap.add_argument("--blank-rows", type=int, default=1, help="Consecutive blank rows to split tables (default: 1)")
    args = ap.parse_args()

    input_root = Path(args.in_dir).resolve()
    output_root = Path(args.out_dir).resolve()

    if not input_root.exists():
        print(
            f'Input directory not found: {input_root}  （repo直下に「設計書」フォルダを作ってxlsx/xls/docxファイルを置いてください）',
            file=sys.stderr,
        )
        return 2

    ok, ng = convert_tree(input_root, output_root, blank_rows=max(1, args.blank_rows))
    print(f"\nDone. documents OK={ok}, NG={ng}")
    return 1 if ng > 0 else 0


if __name__ == "__main__":
    raise SystemExit(main())
